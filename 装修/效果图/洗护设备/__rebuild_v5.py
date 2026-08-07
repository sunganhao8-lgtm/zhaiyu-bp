"""
v5 重建脚本 — 霍曼新样机价 ¥18,299 后重新计算节省
v4 → v5 改动:
- 霍曼基准价: ¥21,978 (7 件) → ¥18,299 (6 件样机不含圆桌) + 圆桌 BT10 ¥720 另买 = 7 件 ¥19,019
- v4 平替价 ¥12,243 不变
- 节省: ¥9,735 (44.3% off) → ¥6,776 (35.6% off 新霍曼 7 件)
- 资金分配不变 (设备实付还是 ¥18,243, 流动资金 ¥48,740, 营销 ¥35,017)
- 启动资金总额 18.23 万 保持不变
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

PATH = r'C:\Users\11390\Desktop\zhaiyu-bp\装修\效果图\洗护设备\设备对比表-模板-v5.xlsx'
wb = openpyxl.load_workbook(PATH, data_only=False)

# ========== Sheet 0-霍曼对话报价 v5 改 ==========
ws0 = wb['0-霍曼对话报价']

# Row 1 标题加 v5
ws0.cell(row=1, column=1, value='霍曼最新报价 · 微信对话原文整理(2026-08-07 v5 新样机价) — 仅作 audit 凭证,不作选品')

# Row 4-11 改新价
# Row 4: 圆桌 (霍曼新方案不含, 720 另算)
ws0.cell(row=4, column=1, value=1)
ws0.cell(row=4, column=2, value='圆桌')
ws0.cell(row=4, column=3, value='—(新样机价方案不含)')
ws0.cell(row=4, column=4, value=720)
ws0.cell(row=4, column=5, value=1)
ws0.cell(row=4, column=6, value=720)
ws0.cell(row=4, column=7, value='对应 BT10(新样机 6 件方案不含,需另买)')

# Row 5: 低压洗护台 LGT110 3300
ws0.cell(row=5, column=1, value=2)
ws0.cell(row=5, column=2, value='低压洗护台')
ws0.cell(row=5, column=3, value='110cm')
ws0.cell(row=5, column=4, value=3300)
ws0.cell(row=5, column=5, value=1)
ws0.cell(row=5, column=6, value=3300)
ws0.cell(row=5, column=7, value='对应 LGT110(原 3680,样机 -380)')

# Row 6: 扫风烘干箱 SD80 2099
ws0.cell(row=6, column=1, value=3)
ws0.cell(row=6, column=2, value='扫风烘干箱')
ws0.cell(row=6, column=3, value='—')
ws0.cell(row=6, column=4, value=2099)
ws0.cell(row=6, column=5, value=1)
ws0.cell(row=6, column=6, value=2099)
ws0.cell(row=6, column=7, value='对应 SD80(原 2399,样机 -300)')

# Row 7: 不锈钢升降浴缸 PB20 4000
ws0.cell(row=7, column=1, value=4)
ws0.cell(row=7, column=2, value='不锈钢升降浴缸')
ws0.cell(row=7, column=3, value='—')
ws0.cell(row=7, column=4, value=4000)
ws0.cell(row=7, column=5, value=1)
ws0.cell(row=7, column=6, value=4000)
ws0.cell(row=7, column=7, value='对应 PB20(原 4620,样机 -620)')

# Row 8: 350 升大烘干箱 PDB350 6500
ws0.cell(row=8, column=1, value=5)
ws0.cell(row=8, column=2, value='350 升大烘干箱')
ws0.cell(row=8, column=3, value='350 升')
ws0.cell(row=8, column=4, value=6500)
ws0.cell(row=8, column=5, value=1)
ws0.cell(row=8, column=6, value=6500)
ws0.cell(row=8, column=7, value='对应 PDB350(原 8000,样机 -1500)')

# Row 9: 升降方桌 BT20M 1800
ws0.cell(row=9, column=1, value=6)
ws0.cell(row=9, column=2, value='升降方桌')
ws0.cell(row=9, column=3, value='—')
ws0.cell(row=9, column=4, value=1800)
ws0.cell(row=9, column=5, value=1)
ws0.cell(row=9, column=6, value=1800)
ws0.cell(row=9, column=7, value='对应 BT20M(原 2100,样机 -300)')

# Row 10: ND10 600(原 459 吹水机赠品,样机方案改为正式件)
ws0.cell(row=10, column=1, value=7)
ws0.cell(row=10, column=2, value='挂脖吹水机')
ws0.cell(row=10, column=3, value='¥600')
ws0.cell(row=10, column=4, value=600)
ws0.cell(row=10, column=5, value=1)
ws0.cell(row=10, column=6, value=600)
ws0.cell(row=10, column=7, value='对应 ND10(原 459 赠品 → 600 样机正式件,+141)')

# Row 11: 不再重复,改为说明
ws0.cell(row=11, column=1, value=None)
ws0.cell(row=11, column=2, value='—(原方案 8 件)新方案 7 件:6 件样机 + 圆桌另买')
ws0.cell(row=11, column=3, value='—')
ws0.cell(row=11, column=4, value=None)
ws0.cell(row=11, column=5, value=None)
ws0.cell(row=11, column=6, value=None)
ws0.cell(row=11, column=7, value='6 件样机价 18,299 + 圆桌 720 = 7 件 19,019(原 21,978 -2,959)')

# Row 12-13 公式保留 (SUM 不变)

print('Sheet 0 改完')

# ========== Sheet 1-霍曼设备主表 v5 改 ==========
ws1 = wb['1-霍曼设备主表']

# K 列价格
# Row 4: ND10 459 → 600
ws1.cell(row=4, column=11, value=600)
# Row 5: BT10 720 (霍曼新方案不含, 单买) — 不变
# Row 6: LGT110 3680 → 3300
ws1.cell(row=6, column=11, value=3300)
# Row 7: SD80 2399 → 2099
ws1.cell(row=7, column=11, value=2099)
# Row 8: PB20 4620 → 4000
ws1.cell(row=8, column=11, value=4000)
# Row 9: PDB350 8000 → 6500
ws1.cell(row=9, column=11, value=6500)
# Row 10: BT20M 2100 → 1800
ws1.cell(row=10, column=11, value=1800)
# Row 11-12: 赠品 — 不变

# Row 13: 套餐合计 K13 公式含 K4-K10 — 公式不变
# 但 v5 是 7 件 = 19,019, v4 是 21,978
# 公式是 =SUM(K4:K10) 自动重算 19,019
# 加备注
ws1.cell(row=14, column=11, value=None)
# 改 Row 14 备注
# 找 Row 14 备注列
print('Sheet 1 改完')

# ========== Sheet 3-决策汇总 v5 改 ==========
ws3 = wb['3-决策汇总']

# Row 1: 标题
ws3.cell(row=1, column=1, value='决策辅助 — 空间适配检查 + 总价对比 + 最终选品 【v5: 霍曼新样机价 18,299(6 件)+ 圆桌另买 = 19,019,v4 平替 12,243 省 6,776 (35.6% off)】')

# Row 18: 方案 A 霍曼全套 改新价
ws3.cell(row=18, column=1, value='方案 A · 霍曼新样机价')
ws3.cell(row=18, column=3, value=19019)  # 6 件样机 18,299 + 圆桌另买 720
ws3.cell(row=18, column=4, value='霍曼 6 件样机(LGT110+SD80+PDB350+PB20+BT20M+ND10)¥18,299 + 圆桌 BT10 ¥720 另买 = ¥19,019(原 ¥21,978 降 ¥2,959)')
ws3.cell(row=18, column=5, value='⭐⭐⭐⭐ 性能完美匹配 PDF / 样机 8-9 成新 / 质保从购买日起算(跟新机一样)')
ws3.cell(row=18, column=7, value='现货 / 展会处理')
ws3.cell(row=18, column=8, value='圆桌不在样机方案需另买(¥720);吹水机从赠品升级为正式件(+¥141)')
ws3.cell(row=18, column=9, value='☞ 稳健首选(¥2,959 折扣但仍比平替贵 ¥6,776)')

# Row 20: 方案 C 全平替 v5 改
ws3.cell(row=20, column=1, value='方案 C · 全平替 【v5 落地】')
ws3.cell(row=20, column=3, value=18243)  # 设备实付 = 12,243 + 6,000 囤货证照
ws3.cell(row=20, column=4, value='v5: v4 平替 7 件 ¥12,243 + 囤货证照 ¥6,000 vs 霍曼新样机 7 件 ¥19,019,省 ¥6,776(35.6% off 新样机)')
ws3.cell(row=20, column=5, value='⭐⭐⭐ 段#3 完美对标 430L 蔚德晟(容量 1.23x+功率 1.27x+内尺寸 1.98x);其余 5 段降本')
ws3.cell(row=20, column=7, value='48h-7 天')
ws3.cell(row=20, column=8, value='v4 平替方案不变;霍曼新样机价降 ¥2,959 但平替仍省 ¥6,776')
ws3.cell(row=20, column=9, value='☞ 已选(平替 35.6% off 新样机)')

# Row 35: 段#3 选品不变 (v4 蔚德晟 430L ¥4880)

# Row 40: 合计行 — 改"7 件平替" 对标"霍曼新样机 7 件 19,019"
ws3.cell(row=40, column=3, value=19019)  # 霍曼新 7 件总价
ws3.cell(row=40, column=4, value='7 件平替组合(段#2 A/B 备份 + 段#3 完美对标 430L 蔚德晟,其他 5 件各 1 个绿候选) vs 霍曼新样机 7 件')
ws3.cell(row=40, column=6, value=12243)
ws3.cell(row=40, column=7, value='35.6%')

# Row 42: 3.5 决策建议 v5
ws3.cell(row=42, column=1, value='3.5 决策建议 【v5: 霍曼新样机价 18,299 + 圆桌 720 = 7 件 19,019;v4 平替 12,243 省 6,776 (35.6%)】')

# Row 43: 主推方案
ws3.cell(row=43, column=1, value='✅ 主推方案 v5:7 件平替 ¥12,243 vs 霍曼新样机 7 件 ¥19,019,节省 ¥6,776 (35.6% off 新样机) — 段#2 A/B 备份(嬉皮狗 ¥603 + 嘉兴慧曼 ¥288) + 段#3 完美对标 430L 蔚德晟 ¥4880 + 段#5/段#6/段#7 维持 v4 选品')

# Row 44: v5 关键发现 (vs v4)
ws3.cell(row=44, column=1, value='🎯 v5 关键发现 (vs v4) — 霍曼新样机价 6 件 ¥18,299,圆桌另买 ¥720:')

# Row 45: 关键发现 1
ws3.cell(row=45, column=1, value='   • 霍曼降价:7 件套餐 ¥21,978 → ¥19,019(降 ¥2,959 / -13.5% off) — 6 件样机(LGT110+SD80+PDB350+PB20+BT20M+ND10)= ¥18,299,圆桌 BT10 不含需另买 ¥720')

# Row 46: 关键发现 2
ws3.cell(row=46, column=1, value='   • 段#3 PDB350 对比:霍曼样机 ¥6,500(原 8,000 降 1,500)vs v4 平替 430L 蔚德晟 ¥4,880 — 平替仍便宜 ¥1,620(24.9% off 新霍曼) + 容量 1.23x + 功率 1.27x 优势保持')

# Row 47: 关键发现 3
ws3.cell(row=47, column=1, value='   • 段#5 BT20M 对比:霍曼样机 ¥1,800(原 2,100 降 300)vs v4 平替佛山酷格 ¥519 — 平替便宜 ¥1,281(71.2% off 新霍曼) — 霍曼新价反而拉大段#5 价差')

# Row 48: 关键发现 4
ws3.cell(row=48, column=1, value='   • 段#7 ND10 吹水机:霍曼样机方案从原赠品 ¥459 升为正式件 ¥600(+¥141, 实际涨价)vs v4 平替江苏超伦 ¥299 — 平替仍便宜 ¥301(50.2% off 新霍曼)')

# Row 49: 资金再分配 v5 (不变 — 设备实付 18,243 沿用 v4)
ws3.cell(row=49, column=1, value='📋 资金再分配 (v5 vs v4):')

# Row 50: 设备实付
ws3.cell(row=50, column=1, value='   • 设备实付:¥18,243 (¥12,243 设备 + ¥6,000 囤货/证照) — v5 沿用 v4 数字(平替价不变)')

# Row 51: 流动资金
ws3.cell(row=51, column=1, value='   • 流动资金:¥48,740 (v4 不变,v5 沿用)')

# Row 52: 营销
ws3.cell(row=52, column=1, value='   • 营销:¥35,017 (v4 不变)')

# Row 53: 启动资金总额
ws3.cell(row=53, column=1, value='   • 启动资金总额:18.23 万(保持不变,仅分配结构与 v4 一致)')

# Row 54: 下一步
ws3.cell(row=54, column=1, value='🚀 下一步:每件拿 1 件样品实物对比 → 段#3 蔚德晟 430L 启动前必须先发样品验证双区独立控制 + 2800W 走单独 16A 插座 → 段#5 酷格黑色通款小启动前核价 → git commit v5 → GitHub Pages 线上版更新')

print('Sheet 3 改完')

# ========== Sheet 4-评分体系 v5 描述 ==========
ws4 = wb['4-评分体系']

# Row 27: 评分分布 v5
ws4.cell(row=27, column=1, value='📊 v5 评分分布(23 候选):🟢绿(≥80)=15 / 🟡黄(60-79)=8 / 🔴红(<60)=0 — 沿用 v4 分布(选品无变化,仅霍曼基准价改 21,978→19,019)')

# Row 28: v5 变化
ws4.cell(row=28, column=1, value='💡 v5 变化:霍曼 7 件套餐 ¥21,978 → 新样机价 ¥19,019(LGT110+SD80+PDB350+PB20+BT20M+ND10 共 6 件 ¥18,299 + 圆桌 ¥720 另买)。v4 平替 7 件 ¥12,243 不变,新节省 ¥6,776 (35.6% off 新霍曼) — 比 v4 节省 ¥9,735 (44.3% off 原价) 少省 ¥2,959(因霍曼新价也降了)。资金分配不变,启动资金 18.23 万不变。')

# Sheet 1 霍曼报价表(Row 4-10 K 列价)+ 备注
ws1.cell(row=2, column=1, value='宅域洗护区 · 霍曼设备清单(基准表 v5 — 2026-08-07 新样机价 ¥19,019 vs v4 选品 12,243;平替时所有品牌都跟这一行的尺寸/性能横向比对)')

# 设备价格备注
# Row 4 ND10 600 备注
# 不改 PDF 来源列 (L 列)
# 改 PDF 来源页 L4 = 8 (维持)
# 改 R4 备注 — 改 霍曼原 459 现 600

# 改 Sheet 1 公式
# R13 公式 =SUM(K4:K10) 已经是公式

# 检查 1 段 - 圆桌 720
# Sheet 1 K5 720 备注保留 — 新方案不含,单买 720

print('Sheet 1+4 改完')

wb.save(PATH)
print(f'\n✅ v5 xlsx 保存: {PATH}')
print(f'7 件霍曼新样机价 = 600+720+3300+2099+4000+6500+1800 = {600+720+3300+2099+4000+6500+1800}')
print(f'vs v4 平替 12,243 → 节省 {600+720+3300+2099+4000+6500+1800-12243} ({(600+720+3300+2099+4000+6500+1800-12243)/(600+720+3300+2099+4000+6500+1800)*100:.1f}% off)')
