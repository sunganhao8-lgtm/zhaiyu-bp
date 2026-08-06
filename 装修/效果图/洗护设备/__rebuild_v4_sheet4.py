"""
补 Sheet 4 v4 描述
"""
import openpyxl

PATH = r'C:\Users\11390\Desktop\zhaiyu-bp\装修\效果图\洗护设备\设备对比表-模板-v4.xlsx'
wb = openpyxl.load_workbook(PATH, data_only=False)
ws4 = wb['4-评分体系']

# Row 27: 评分分布改 v4
ws4.cell(row=27, column=1, value='📊 v4 评分分布(23 候选):🟢绿(≥80)=15 / 🟡黄(60-79)=8 / 🔴红(<60)=0')

# Row 28: v4 变化
ws4.cell(row=28, column=1, value='💡 v4 变化:段#3 致命修复 — 75L 顺旺通达(0.21x 容量 红灯)替换为 430L 蔚德晟 ¥4880(1.23x 容量 绿灯 95 分)。其他 6 段维持 v3 选品(段#1/段#2/段#4/段#5/段#6/段#7)。')

wb.save(PATH)
print('✅ Sheet 4 v4 描述更新完成')
