"""
v4 重建脚本 — 应用 v4 精修到 xlsx
v3 → v4 改动:
- 段#3 烘干箱大 致命问题: 75L 顺旺通达 ¥749 → 430L 蔚德晟 ¥4880 (完美对标 350L 霍曼)
- 其他 6 段维持 v3
- 资金再分配: 设备实付 +¥4,131, 流动资金 -¥4,131, 营销不变
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = r'C:\Users\11390\Desktop\zhaiyu-bp\装修\效果图\洗护设备\设备对比表-模板-v4.xlsx'
wb = openpyxl.load_workbook(PATH, data_only=False)

# ========== Sheet 2 段#3 改选 ==========
ws2 = wb['2-平替对比']

# Row 16 改为: 主选 蔚德晟 430L ¥4880
row16_values = [
    None,
    '东莞蔚德晟 430L 超巨大可折叠(对标霍曼 PDB350) 【v4 主选 完美对标】',
    'A',
    '东莞市蔚德晟科技有限公司 430L 超巨大可折叠烘干箱',
    '100L ¥1880 / 200L ¥2580 / 430L ¥4880 / 120L(40斤) ¥1380 / 120L(60斤) ¥1899',
    4880,
    '430L 超巨大可折叠 100×72×96.5cm / 55kg / 2800W / 110斤 (双区独立控制+臭氧消毒+负离子)',
    '220V / 2800W / 直流电机 / 波纹丝 / 双区独立控制 / 臭氧杀菌 / 负离子 / 跨境资质',
    '≤55kg 大型犬(金毛/边牧/两犬同洗)',
    '220V / 跨境出口专供',
    '12 月',
    '广东东莞 / 1 个起批 / 满 1000 包邮 / 库存 10000',
    'https://detail.1688.com/offer/965079465804.html',
    '东莞市蔚德晟科技有限公司(主营猫/狗美容清洁用品)',
    '回头率 40% / 服务 4.0 / 好评 100% / 100+ 件已售',
    '[绿灯 95分] ¥20/材20/规20/店12/交23 — 容量 430L vs 霍曼 350L = 1.23x / 功率 2800W vs 2200W = 1.27x / 内尺寸 1.98x / 双区独立控制+臭氧+负离子 / 价格省 39%',
    95,
    '✅ 已选(主用 完美对标)',
]
for col, val in enumerate(row16_values, 1):
    ws2.cell(row=16, column=col, value=val)

# Row 17 改为: 备选 蔚德晟 200L ¥2580
row17_values = [
    None,
    '东莞蔚德晟 200L 超大容量烘干箱(对标霍曼 PDB350) 【v4 备选】',
    'B',
    '东莞市蔚德晟科技有限公司 200L 超大容量烘干箱',
    '200L 66.5×72×78.5cm / 32kg / 1400W / 30-50斤 猫狗实用',
    2580,
    '200L 超大容量 66.5×72×78.5cm / 32kg / 1400W / 30-50斤',
    '220V / 1400W / 直流电机 / 波纹丝 / 臭氧杀菌+自动集毛+负离子+猫狗窝',
    '≤25kg 中大型犬',
    '220V / 跨境出口专供',
    '12 月',
    '广东东莞 / 1 个起批 / 库存 10000',
    'https://detail.1688.com/offer/965079465804.html',
    '东莞市蔚德晟科技有限公司',
    '回头率 40% / 服务 4.0 / 好评 100%',
    '[黄灯 76分] ¥18/材18/规14/店12/交14 — 容量 200L vs 350L = 0.57x / 适用 25kg 偏小 / 1400W / 价格省 68%',
    76,
    '⚠ 备选(容量偏小,只适合 ≤25kg)',
]
for col, val in enumerate(row17_values, 1):
    ws2.cell(row=17, column=col, value=val)

# Row 18 改为: 备选 蔚德晟 120L 60斤 ¥1899
row18_values = [
    None,
    '东莞蔚德晟 120L 宠有家(60斤以内)(对标霍曼 PDB350) 【v4 备选】',
    'C',
    '东莞市蔚德晟科技有限公司 120L 宠有家(60斤以内)烘干箱',
    '120L 65×60×73cm / 37kg / 60斤(30kg)以内 / 标题虚标实测约 285L',
    1899,
    '120L 65×60×73cm / 37kg / 60斤(30kg)以内 (标题虚标,实测约 285L)',
    '220V / 直流电机 / 跨境资质',
    '≤30kg 中大型犬',
    '220V / 跨境出口专供',
    '12 月',
    '广东东莞 / 1 个起批 / 库存 10000',
    'https://detail.1688.com/offer/965079465804.html',
    '东莞市蔚德晟科技有限公司',
    '回头率 40% / 服务 4.0 / 好评 100%',
    '[黄灯 70分] ¥18/材18/规12/店12/交10 — 标题 120L 实测 285L / 适用 30kg 偏小 / 价格省 76%',
    70,
    '⚠ 备选(容量小 + 适用偏轻)',
]
for col, val in enumerate(row18_values, 1):
    ws2.cell(row=18, column=col, value=val)

# 更新 Row 2 摘要
ws2.cell(row=2, column=1, value='✅ v4 已填,#1(3)/#2(3)/#3(3)/#4(4)/#5(4)/#6(3)/#7(3) — 共 23 候选。🟢绿 15 / 🟡黄 8 / 🔴红 0。【v4 改动】段#3 烘干箱大 致命问题修复:75L 顺旺通达 ¥749 → 430L 蔚德晟 ¥4880(完美对标 350L 霍曼:容量 1.23x + 功率 1.27x + 内尺寸 1.98x + 价格省 39%)。')
print('Sheet 2 段#3 改选完成')

# ========== Sheet 3 段#3 改选 + 合计改 ¥12,243 ==========
ws3 = wb['3-决策汇总']

# Row 1: 标题
ws3.cell(row=1, column=1, value='决策辅助 — 空间适配检查 + 总价对比 + 最终选品 【v4: 段#3 完美对标 350L 霍曼 PDB350】')

# Row 8: PDB350 空间适配不变(尺寸 105×73×98 不变)

# Row 20: 方案 C 改 v4
ws3.cell(row=20, column=1, value='方案 C · 全平替 【v4 落地】')
ws3.cell(row=20, column=3, value=18243)
ws3.cell(row=20, column=4, value='v4: 7 件平替组合(段#2 A/B 备份 + 段#3 完美对标 430L 蔚德晟)+ 囤货证照 ¥6000')
ws3.cell(row=20, column=5, value='⭐⭐⭐ 段#3 完美对标(430L vs 350L 霍曼)/其余段降本')
ws3.cell(row=20, column=7, value='48h-7 天')
ws3.cell(row=20, column=8, value='段#3 容量差 4.7 倍致命问题已修复;段#2 备份稳;430L 大型箱走单独 16A 插座')
ws3.cell(row=20, column=9, value='☞ 已选(段#3 完美对标 + 总体降本 44.3%)')

# Row 35: 段#3 选品改 蔚德晟 430L
ws3.cell(row=35, column=2, value='PDB350')
ws3.cell(row=35, column=3, value=8000)
ws3.cell(row=35, column=4, value='430L 超巨大可折叠 100×72×96.5cm / 55kg / 2800W / 110斤 (双区独立控制+臭氧+负离子) / 容量 1.23x + 功率 1.27x + 内尺寸 1.98x vs 霍曼 350L')
ws3.cell(row=35, column=5, value='东莞蔚德晟科技')
ws3.cell(row=35, column=6, value=4880)
ws3.cell(row=35, column=7, value='39%')
ws3.cell(row=35, column=8, value='https://detail.1688.com/offer/965079465804.html')

# Row 40: 合计改 ¥12,243 / 44.3%
ws3.cell(row=40, column=3, value=21978)
ws3.cell(row=40, column=4, value='7 件平替组合(段#2 A/B 备份 + 段#3 完美对标 430L 蔚德晟,其他 5 件各 1 个绿候选)')
ws3.cell(row=40, column=6, value=12243)
ws3.cell(row=40, column=7, value='44.3%')

# Row 42: 3.5 决策建议
ws3.cell(row=42, column=1, value='3.5 决策建议 【v4: 段#3 完美对标】')

# Row 43: 主推方案
ws3.cell(row=43, column=1, value='✅ 主推方案 v4:7 件平替 ¥12,243 vs 霍曼 ¥21,978,节省 ¥9,735 (44.3%) — 段#2 A/B 备份(嬉皮狗 ¥603 + 嘉兴慧曼 ¥288) + 段#3 完美对标 430L 蔚德晟 ¥4880(替代 v3 75L 顺旺通达致命错)')

# Row 44: v4 关键发现 (vs v3)
ws3.cell(row=44, column=1, value='🎯 v4 关键发现 (vs v3) — 修复段#3 致命对标偏差:')

# Row 45: 关键发现 1
ws3.cell(row=45, column=1, value='   • 段#3 致命: 75L 顺旺通达 ¥749 vs 350L 霍曼 PDB350 = 容量差 4.7 倍(只能放 ≤25kg 宠物,霍曼 60kg 大型犬) → v4 改 430L 蔚德晟 ¥4880,容量 1.23x + 功率 1.27x + 内尺寸 1.98x + 价格省 39%(双区独立控制+臭氧+负离子)')

# Row 46: 关键发现 2
ws3.cell(row=46, column=1, value='   • 段#1 维持 v3 1.27m 步梯款 ¥2850(偏长 17cm 轻微,v3 已接受;1.2m 电动升降款 ¥3669 更贴但加 ¥819,v4 维持原选)')

# Row 47: 关键发现 3
ws3.cell(row=47, column=1, value='   • 段#2 维持 v3 A/B 备份(嬉皮狗 ¥603 + 嘉兴慧曼 ¥288 = ¥891);1000W vs 霍曼 1200W 差 17% 轻微,1688 市场 ≤12.5kg 烘干箱功率上限 1000W,京东志高 1500W ¥458 可作第三备份')

# Row 48: 关键发现 4
ws3.cell(row=48, column=1, value='   • 段#5 维持 v3 佛山酷格 124×64 ¥519(偏长 28cm 轻微,17 款 SKU 中无 96×65 款;¥519 = 黑色通款小(尺寸未明)需启动前核价)')

# Row 49: 资金再分配
ws3.cell(row=49, column=1, value='📋 资金再分配 (v4 vs v3):')

# Row 50: 设备实付
ws3.cell(row=50, column=1, value='   • 设备实付:¥18,243 (¥12,243 设备 + ¥6,000 囤货/证照) — 比 v3 14,112 多 ¥4,131(段#3 ¥749 → ¥4880)')

# Row 51: 流动资金
ws3.cell(row=51, column=1, value='   • 流动资金:¥48,740 (v3 是 52,871) — 减 ¥4,131')

# Row 52: 营销
ws3.cell(row=52, column=1, value='   • 营销:¥35,017 (不变)')

# Row 53: 启动资金总额
ws3.cell(row=53, column=1, value='   • 启动资金总额:18.23 万(保持不变,仅分配结构调整)')

# Row 54: 下一步
ws3.cell(row=54, column=1, value='🚀 下一步:每件拿 1 件样品实物对比 → 段#3 蔚德晟 430L 启动前必须先发样品验证双区独立控制 + 2800W 走单独 16A 插座 → 段#5 酷格黑色通款小启动前核价 → git commit v4 → GitHub Pages 线上版更新')

print('Sheet 3 决策汇总 v4 改完')

# ========== Sheet 4 评分体系 v4 描述 ==========
ws4 = wb['4-评分体系']
# 找含 'v3 评分分布' 的行
for i, row in enumerate(ws4.iter_rows(values_only=True), 1):
    text = ' | '.join([str(c)[:80] if c else '' for c in row[:6]])
    if any(k in text for k in ['v3 评分', 'v4 评分', 'v3 变化', 'v4 变化', '14 /', '15 /', '8 /', '9 /', '黄灯', '绿灯']):
        print(f"Sheet 4 Row {i}: {text[:200]}")

print('\n=== Sheet 4 Row 22-28 ===')
for i in range(22, 29):
    row = list(ws4.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    text = ' | '.join([str(c)[:100] if c else '' for c in row[:6]])
    if text.strip(' |'):
        print(f"Row {i}: {text[:200]}")

wb.save(PATH)
print(f'\n✅ v4 xlsx 保存: {PATH}')
